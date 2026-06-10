import io
import os
from django.conf import settings
from django.http import HttpResponse
from datetime import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    
    @staticmethod
    def safe_str(value):
        """Безпечне перетворення в рядок"""
        if value is None:
            return ""
        return str(value)
    
    @staticmethod
    def export_participants(conference, submissions):
        """
        Експорт учасників конференції у форматі Excel
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Учасники конференції"
        
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = [
            "№", "ПІБ", "Email", "Роль", "Статус тези",
            "Назва тези", "ID тези", "Дата подання", "Коментар рецензента"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Збираємо унікальних учасників
        participants = {}
        
        for submission in submissions:
            author = submission.author
            
            if author.id not in participants:
                participants[author.id] = {
                    'user': author,
                    'submissions': []
                }
            
            participants[author.id]['submissions'].append(submission)
        
        # Заповнюємо дані
        row = 2
        for author_id, data in participants.items():
            author = data['user']
            user_submissions = data['submissions']
            
            for idx, submission in enumerate(user_submissions):
                # Визначаємо роль
                if submission.author == conference.organizer:
                    role = "Організатор"
                elif submission.reviewer:
                    role = "Рецензент"
                else:
                    role = "Автор"
                
                ws.cell(row=row, column=1, value=row-1).border = border
                ws.cell(row=row, column=2, value=ExcelExporter.safe_str(author.full_name or author.username)).border = border
                ws.cell(row=row, column=3, value=ExcelExporter.safe_str(author.email)).border = border
                ws.cell(row=row, column=4, value=role).border = border
                
                status_cell = ws.cell(row=row, column=5, value=submission.get_status_display())
                status_cell.border = border
                
                if submission.status == 'ACCEPTED':
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif submission.status == 'REJECTED':
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif submission.status == 'PENDING':
                    status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                
                ws.cell(row=row, column=6, value=ExcelExporter.safe_str(submission.title)).border = border
                ws.cell(row=row, column=7, value=submission.id).border = border
                
                submitted_at = ""
                if submission.created_at:
                    submitted_at = submission.created_at.strftime("%d.%m.%Y %H:%M")
                ws.cell(row=row, column=8, value=submitted_at).border = border
                
                ws.cell(row=row, column=9, value=ExcelExporter.safe_str(submission.reviewer_comment or "")).border = border
                
                row += 1
        
        # Якщо немає жодного учасника
        if row == 2:
            ws.cell(row=2, column=1, value="Немає учасників").border = border
        
        # Автоматичне налаштування ширини колонок
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for r in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=r, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Додаємо інформацію внизу
        info_row = ws.max_row + 2
        ws.cell(row=info_row, column=1, value=f"Конференція: {ExcelExporter.safe_str(conference.title)}")
        ws.cell(row=info_row + 1, column=1, value=f"ID конференції: {ExcelExporter.safe_str(conference.conference_id)}")
        ws.cell(row=info_row + 2, column=1, value=f"Дата експорту: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        ws.cell(row=info_row + 3, column=1, value=f"Всього учасників: {len(participants)}")
        ws.cell(row=info_row + 4, column=1, value=f"Всього тез: {submissions.count()}")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_submissions_list(conference, submissions):
        """
        Експорт списку тез конференції у форматі Excel
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Тези конференції"
        
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = [
            "№", "ID тези", "Назва тези", "Автор", "Email автора",
            "Статус", "Дата подання", "Рецензент", "Коментар рецензента", "Версія"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        for row, submission in enumerate(submissions, 2):
            ws.cell(row=row, column=1, value=row-1).border = border
            ws.cell(row=row, column=2, value=submission.id).border = border
            ws.cell(row=row, column=3, value=ExcelExporter.safe_str(submission.title)).border = border
            ws.cell(row=row, column=4, value=ExcelExporter.safe_str(submission.author.full_name or submission.author.username)).border = border
            ws.cell(row=row, column=5, value=ExcelExporter.safe_str(submission.author.email)).border = border
            
            status_cell = ws.cell(row=row, column=6, value=submission.get_status_display())
            status_cell.border = border
            
            if submission.status == 'ACCEPTED':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif submission.status == 'REJECTED':
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif submission.status == 'PENDING':
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif submission.status == 'REVISION_REQUIRED':
                status_cell.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
            
            submitted_at = ""
            if submission.created_at:
                submitted_at = submission.created_at.strftime("%d.%m.%Y %H:%M")
            ws.cell(row=row, column=7, value=submitted_at).border = border
            
            reviewer_name = ""
            if submission.reviewer:
                reviewer_name = submission.reviewer.full_name or submission.reviewer.username
            ws.cell(row=row, column=8, value=ExcelExporter.safe_str(reviewer_name)).border = border
            
            ws.cell(row=row, column=9, value=ExcelExporter.safe_str(submission.reviewer_comment or "")).border = border
            ws.cell(row=row, column=10, value=submission.version).border = border
        
        # Якщо немає тез
        if submissions.count() == 0:
            ws.cell(row=2, column=1, value="Немає поданих тез").border = border
        
        # Автоматичне налаштування ширини колонок
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for r in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=r, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Додаємо інформацію внизу
        info_row = ws.max_row + 2
        ws.cell(row=info_row, column=1, value=f"Конференція: {ExcelExporter.safe_str(conference.title)}")
        ws.cell(row=info_row + 1, column=1, value=f"ID конференції: {ExcelExporter.safe_str(conference.conference_id)}")
        ws.cell(row=info_row + 2, column=1, value=f"Дата експорту: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        ws.cell(row=info_row + 3, column=1, value=f"Всього тез: {submissions.count()}")
        ws.cell(row=info_row + 4, column=1, value=f"Прийнято: {submissions.filter(status='ACCEPTED').count()}")
        ws.cell(row=info_row + 5, column=1, value=f"На рецензуванні: {submissions.filter(status='PENDING').count()}")
        ws.cell(row=info_row + 6, column=1, value=f"Відхилено: {submissions.filter(status='REJECTED').count()}")
        ws.cell(row=info_row + 7, column=1, value=f"Потребують доопрацювання: {submissions.filter(status='REVISION_REQUIRED').count()}")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

class PDFExporter:
    FONTS_DIR = os.path.join(os.path.dirname(__file__), 'fonts')
    
    _fonts_registered = False
    
    @classmethod
    def register_fonts(cls):
        if cls._fonts_registered:
            return True
        
        if not os.path.exists(cls.FONTS_DIR):
            print(f"Папка зі шрифтами не знайдена: {cls.FONTS_DIR}")
            return False
        
        regular_path = os.path.join(cls.FONTS_DIR, 'DejaVuSans.ttf')
        bold_path = os.path.join(cls.FONTS_DIR, 'DejaVuSans-Bold.ttf')
        
        try:
            if os.path.exists(regular_path):
                pdfmetrics.registerFont(TTFont('DejaVu', regular_path))
                cls._fonts_registered = True
            
            if os.path.exists(bold_path):
                pdfmetrics.registerFont(TTFont('DejaVu-Bold', bold_path))
            
            pdfmetrics.registerFontFamily(
                'DejaVu',
                normal='DejaVu',
                bold='DejaVu-Bold'
            )
            
        except Exception as e:
            print(f"Помилка реєстрації шрифтів: {e}")
            return False
        
        return cls._fonts_registered
    
    @classmethod
    def export_submission_pdf(cls, submission):
        has_cyrillic_font = cls.register_fonts()
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="thesis_{submission.id}.pdf"'
        
        doc = SimpleDocTemplate(
            response,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        if has_cyrillic_font:
            base_font = 'DejaVu'
            base_font_bold = 'DejaVu-Bold'
        else:
            base_font = 'Helvetica'
            base_font_bold = 'Helvetica-Bold'
        
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Normal'],
            fontName=base_font_bold,
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            alignment=TA_CENTER,
            spaceAfter=30,
            leading=22
        )
        
        section_style = ParagraphStyle(
            'Section',
            parent=styles['Normal'],
            fontName=base_font_bold,
            fontSize=14,
            textColor=colors.HexColor('#366092'),
            spaceBefore=20,
            spaceAfter=10,
            leading=18
        )
        
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontName=base_font,
            fontSize=11,
            spaceAfter=12,
            leading=14,
            alignment=TA_JUSTIFY
        )
        
        table_style = ParagraphStyle(
            'TableStyle',
            parent=styles['Normal'],
            fontName=base_font,
            fontSize=10,
            leading=12
        )
        
        status_colors = {
            'ACCEPTED': colors.green,
            'REJECTED': colors.red,
            'PENDING': colors.orange,
            'REVISION_REQUIRED': colors.orange,
            'DRAFT': colors.grey,
        }
        status_color = status_colors.get(submission.status, colors.black)
        
        status_style = ParagraphStyle(
            'Status',
            parent=normal_style,
            fontName=base_font_bold,
            fontSize=12,
            textColor=status_color,
            alignment=TA_LEFT,
            spaceAfter=12
        )
        
        footer_style = ParagraphStyle(
            'Footer',
            parent=normal_style,
            fontName=base_font,
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )
        
        def safe_paragraph(text):
            if text is None:
                text = ""
            text = str(text)
            text = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            return Paragraph(text, table_style)
        
        story = []
        
        story.append(Paragraph(f"Теза: {submission.title}", title_style))
        
        story.append(Paragraph("Інформація про конференцію", section_style))
        
        info_data = [
            [safe_paragraph("Назва:"), safe_paragraph(submission.conference.title)],
            [safe_paragraph("Категорія:"), safe_paragraph(submission.conference.get_category_display())],
            [safe_paragraph("Дата проведення:"), safe_paragraph(submission.conference.event_date.strftime("%d.%m.%Y"))],
            [safe_paragraph("Дедлайн подачі:"), safe_paragraph(submission.conference.submission_deadline.strftime("%d.%m.%Y"))],
        ]
        
        info_table = Table(info_data, colWidths=[4*cm, 10*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), base_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        story.append(Paragraph("Автори", section_style))
        story.append(Paragraph(submission.author.full_name, normal_style))
        
        if submission.abstract:
            story.append(Paragraph("Анотація", section_style))
            story.append(Paragraph(submission.abstract, normal_style))
        
        story.append(Paragraph("Статус", section_style))
        status_text = dict(submission.STATUS_CHOICES).get(submission.status, submission.status)
        story.append(Paragraph(status_text, status_style))
        
        if submission.reviewer_comment:
            story.append(Paragraph("Коментар рецензента", section_style))
            story.append(Paragraph(submission.reviewer_comment, normal_style))
        
        story.append(Paragraph("Метадані", section_style))
        meta_data = [
            [safe_paragraph("Дата подання:"), safe_paragraph(submission.created_at.strftime("%d.%m.%Y %H:%M"))],
            [safe_paragraph("Останнє оновлення:"), safe_paragraph(submission.updated_at.strftime("%d.%m.%Y %H:%M"))],
            [safe_paragraph("Версія:"), safe_paragraph(str(submission.version))],
            [safe_paragraph("Переглядів:"), safe_paragraph(str(submission.views_count))],
        ]
        
        meta_table = Table(meta_data, colWidths=[4*cm, 10*cm])
        meta_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), base_font),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(meta_table)
        
        story.append(Spacer(1, 30))
        story.append(Paragraph(f"Експортовано: {datetime.now().strftime('%d.%m.%Y %H:%M')}", footer_style))
        
        doc.build(story)
        
        return response